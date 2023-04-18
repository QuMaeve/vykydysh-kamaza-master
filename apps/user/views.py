from django.shortcuts import render, redirect

from django.contrib.auth.decorators import permission_required
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import models
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.db.models import Q
from datetime import date
from dateutil.relativedelta import *

from apps.user.models import CustomUser
from apps.studentclass.models import Class
from apps.establishment.models import Establishment
from django.contrib.auth.models import Group
from apps.main.forms import ImportForm
# from apps import class
# Import `xlrd`
# import xlrd
import re 
from openpyxl import load_workbook
from django.contrib.auth.hashers import make_password
import xlwt
from django.http import HttpResponse
from dateutil.relativedelta import *
import datetime
from django.utils import timezone


from apps.user.forms import CustomUserCreationForm
# Create your views here.

@login_required(login_url='login')
@permission_required('customuser.add_customuser', raise_exception=True)
def create(request):
    form = CustomUserCreationForm()
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Запись успешно создана')
            return redirect('user_index')
        else:
            messages.error(request, 'Введите корректные данные')
    return render(request, 'pages/admin/user_create.html', {'form': form})

@login_required(login_url='login')
def index(request):
    form = ImportForm()
    g_name='Ученики'
    return render(
        request, 
        'pages/admin/import.html', 
        {'form':form,'g_name':g_name,}
        )

@login_required(login_url='login')
def user_index(request):
    data=CustomUser.objects.all()
    return render(
        request, 
        'pages/admin/user.html', 
        {'data':data,}
        )

@login_required(login_url='login')
def index_teacher(request):
    form = ImportForm()
    g_name='Преподаватель'
    return render(
        request, 
        'pages/admin/import_teacher.html', 
        {'form':form,'g_name':g_name,}
        )

@login_required(login_url='login')
def import_user(request):
    form = ImportForm()
    if request.method == 'POST' and request.FILES:
        print(request.POST['g_name'])
        g_name=request.POST['g_name']
        print(request.FILES['file'])
        # Open a workbook 
        # ворк wb = load_workbook('documents/экспорт учащихся.xlsx')
        wb = load_workbook(filename=request.FILES['file'].file)
        sheet = wb.active
        # Retrieve cell value 
        sheet.cell(row=1, column=2).value
        school = Establishment.objects.filter(name=sheet.cell(row=1, column=1).value).first() 
        row_count = sheet.max_row+1
        print("row_count:"+str(row_count))
        user_list=[]
        for i in range(11, row_count):
            print(i)
            if sheet.cell(row=i, column=1).value != None:
                classes = sheet.cell(row=i, column=1).value
                result = "".join(re.findall(r'\d+', classes))
                print(result)
                litera=""
                for d in classes:
                    if d.isalpha():
                        litera = "".join(d)
                if litera !="":
                    print(litera)
                if Class.objects.filter(number=int(result), 
                                        literature = litera,
                                        establishment = 
                                        Establishment.objects.filter(name=sheet.cell(row=1, column=1).value).first(),
                                        ).exists():
                    a =Class.objects.filter(number=int(result), 
                                        literature = litera,
                                        establishment = 
                                            Establishment.objects.filter(name=sheet.cell(row=1, column=1).value).first(),
                                        ).first()
                    print(a)
                else:
                    a = Class.objects.create(number=int(result), 
                                        literature = litera,
                                        establishment = 
                                        Establishment.objects.filter(name=sheet.cell(row=1, column=1).value).first())
                    print('Создан новый класс')
                    print(Class.objects.get(id=a.id))
            first_name=''
            if sheet.cell(row=i, column=2).value != None:
                first_name = sheet.cell(row=i, column=2).value
                print(first_name)
            last_name=''
            if sheet.cell(row=i, column=3).value != None:
                last_name = sheet.cell(row=i, column=3).value
                print(last_name)
            patronymic=''   
            if sheet.cell(row=i, column=4).value != None:
                patronymic = sheet.cell(row=i, column=4).value
                print(patronymic) 
            login = first_name+'_'+last_name[0]+patronymic[0]
            print('login',login)
            for user in CustomUser.objects.all():
                if user.username == login:
                    login=login+'1'
            date_joined = date.today().strftime("%Y-%m-%d")
            group = Group.objects.filter(name=g_name).first()
            random_password = CustomUser.objects.make_random_password()

            u = CustomUser.objects.create(username = login, 
                                    password = make_password(random_password, 'qwertyuiop', 'sha1'),
                                    cleaned_password = random_password,
                                    is_superuser = False,
                                    date_joined = date_joined,
                                    is_active = True,
                                    is_staff = False,
                                    first_name=first_name,
                                    last_name=last_name,
                                    patronymic=patronymic, 
                                    classes=a, 
                                    establishment=school, )
            print(login+' add')
            u.groups.add(group)
            user_list.append(u)
            # Cleaned_password to excel
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f'attachment; filename="report_user_{datetime.date.today().strftime("%Y-%m-%d")}.xls"'
        
        wb = xlwt.Workbook(encoding='cp-1251')

        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        columns = [
            'Логин',
            'Пароль',
        ]
        ws = wb.add_sheet('Sheet1')
        row_num = 0
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
            
        ws.col(0).width = 256 * 36
        ws.col(1).width = 256 * 36
        font_style = xlwt.XFStyle()
        for user in user_list:
            row_num += 1
            col_num = 0
            style = xlwt.easyxf(f'pattern: pattern solid, fore_colour white')
            for value in {
                "Логин": user.username,
                "Пароль": user.cleaned_password,
                }.values():
                    ws.write(row_num, col_num, value, style)
                    col_num += 1
        row_num += 1
        col_num = 0
        messages.success(request, 'Файл сформирован')
        wb.save(response)
        return response
    return redirect('user')

@login_required(login_url='login')
def import_teacher(request):
    form = ImportForm()
    if request.method == 'POST' and request.FILES:
        print(request.POST['g_name'])
        g_name=request.POST['g_name']
        print(request.FILES['file'])
        # Open a workbook 
        # ворк wb = load_workbook('documents/экспорт учащихся.xlsx')
        wb = load_workbook(filename=request.FILES['file'].file)
        sheet = wb.active
        # Retrieve cell value 
        sheet.cell(row=1, column=2).value
        school = Establishment.objects.filter(name=sheet.cell(row=1, column=1).value).first() 
        row_count = sheet.max_row+1
        print("row_count:"+str(row_count))
        user_list=[]
        for i in range(8, row_count):
            print(i)
            first_name=''
            last_name=''
            patronymic=''
            if sheet.cell(row=i, column=2).value != None:
                first_name = str(sheet.cell(row=i, column=2).value).split()[0]
                print(first_name)
                last_name =  str(sheet.cell(row=i, column=2).value).split()[1]
                print(last_name)
                print(len(str(sheet.cell(row=i, column=2).value).split()))
                if len(str(sheet.cell(row=i, column=2).value).split())>2:
                    patronymic =  str(sheet.cell(row=i, column=2).value).split()[2]
                    print(patronymic)
                    login = first_name+'_'+last_name[0]+patronymic[0]
                else:
                    login = first_name+'_'+last_name[0]
                    print('login',login)
                for user in CustomUser.objects.all():
                    if user.username == login:
                        login=login+'1'
                date_joined = date.today().strftime("%Y-%m-%d")
                group = Group.objects.filter(name=g_name).first()
                random_password = CustomUser.objects.make_random_password()

                u = CustomUser.objects.create(username = login, 
                                        password = make_password(random_password, 'qwertyuiop', 'sha1'),
                                        cleaned_password = random_password,
                                        is_superuser = False,
                                        date_joined = date_joined,
                                        is_active = True,
                                        is_staff = False,
                                        first_name=first_name,
                                        last_name=last_name,
                                        patronymic=patronymic, 
                                        establishment=school, )
                print(login+' add')
                u.groups.add(group)
                user_list.append(u)
        # Cleaned_password to excel
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f'attachment; filename="report_user_{datetime.date.today().strftime("%Y-%m-%d")}.xls"'
        
        wb = xlwt.Workbook(encoding='cp-1251')

        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        columns = [
            'Логин',
            'Пароль',
        ]
        ws = wb.add_sheet('Sheet1')
        row_num = 0
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
            
        ws.col(0).width = 256 * 36
        ws.col(1).width = 256 * 36
        font_style = xlwt.XFStyle()
        for user in user_list:
            row_num += 1
            col_num = 0
            style = xlwt.easyxf(f'pattern: pattern solid, fore_colour white')
            for value in {
                "Логин": user.username,
                "Пароль": user.cleaned_password,
                }.values():
                    ws.write(row_num, col_num, value, style)
                    col_num += 1
        row_num += 1
        col_num = 0
        messages.success(request, 'Файл сформирован')
        wb.save(response)
        return response
    return redirect('user')